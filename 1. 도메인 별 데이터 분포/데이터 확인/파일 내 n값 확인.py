import glob
import pandas as pd
from openpyxl import Workbook

# path : 라이프_L, 산업_I, 스포츠_S, 오피니언_O, 연예_E, 증권_T, 부동산_R
#path = "C:\\Users\\이성재\\OneDrive - 데이터누리\회사\\3. 2023 사업\\4. AI 학습용데이터\\01. 한국어_다중이벤트\\13.1 데이터\\1. JSON 데이터\\final\\이벤트_재전송_231123\\이벤트_통폐합\\trainning\\라이프_L\\*"
#path = "C:\\Users\\이성재\\OneDrive - 데이터누리\\회사\\3. 2023 사업\\4. AI 학습용데이터\\01. 한국어_다중이벤트\\13.1 데이터\\1. JSON 데이터\\final\\이벤트_통폐합\\이벤트_통폐합\\home\\13-2MEE\\부동산_R\\*"
#path = 'C:\\Users\\이성재\\OneDrive - 데이터누리\\회사\\3. 2023 사업\\4. AI 학습용데이터\\01. 한국어_다중이벤트\\13.1 데이터\\1. JSON 데이터\\final\\이벤트_354건_후처리_231206\\이벤트_통폐합_1206_복사본\\trainning\\라이프_L\\*'
#path = "C:\\Users\\이성재\\OneDrive - 데이터누리\\바탕 화면\\home\\13-2MEE\\MEE_Trainning\\MEE_Trainning_D_13_L\\*"     # 연예_E, 스포츠_S, 산업_I, 오피니언_O, 부동산_R, 증권_T, 라이프_L
#path = "C:\\Users\\이성재\\OneDrive - 데이터누리\\바탕 화면\\이벤트_통폐합_1206\\증권_T\\*"     # 연예_E, 스포츠_S, 산업_I, 오피니언_O, 부동산_R, 증권_T, 라이프_L
path = "C:\\Users\\이성재\\OneDrive - 데이터누리\\회사\\3. 2023 사업\\4. AI 학습용데이터\\01. 한국어_다중이벤트\\01. 이벤트\\05. 이벤트 통폐합\\1. 통폐합_final\\한국어_최종_제외파일_이었던것\\라이프\\*"
json_file = glob.glob(f'{path}.json')       # 경로 내 JSON파일 잡아냄
len(json_file)

vin = []

for name in json_file:
    with open(name, 'r', encoding='utf-8') as f:
        data_str = f.read()
        
        if '\\n' in data_str:
            print(f'The File {name} contains \\n')
            vin.append(name)
            
print(vin)
vin[1][133:166]
len(vin)

vin2 = []
for i in vin:
    print(i[133:166])
    vin2.append(i[133:166])

data = pd.DataFrame(vin2)

data
data.to_excel("부동산 n존재 파일명.xlsx")            # 부동산_R



# excel
wb = Workbook()
ws = wb.active

ws1 = wb.create_sheet("라이프_L")       # 끝에 삽입(기본값)
ws2 = wb.create_sheet("산업_I", 0)      # 첫 번째 위치에 삽입
ws3 = wb.create_sheet("스포츠_S", -1)   # 끝에서 두 번째 위치에 삽입

wb.sheetnames

ws.title = '오피니언_O'
wb.sheetnames

for sheet in wb:
    print(sheet.title)
    
# save
wb.save("라이프_n값.xlsx")